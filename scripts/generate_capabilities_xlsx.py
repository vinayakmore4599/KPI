#!/usr/bin/env python3
"""Generate an Excel catalog of KPI Engine capabilities for authors and architects."""

from __future__ import annotations

from pathlib import Path

import yaml
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.worksheet.worksheet import Worksheet

ROOT = Path(__file__).resolve().parents[1]
REG = ROOT / "kpi_engine/registries"
OUT = ROOT / "docs/KPI-Engine-Capabilities.xlsx"

NAVY = "1A365D"
GOLD = "C9A227"
LIGHT = "F4F7FB"
WHITE = "FFFFFF"
THIN = Border(
    left=Side(style="thin", color="D0D7DE"),
    right=Side(style="thin", color="D0D7DE"),
    top=Side(style="thin", color="D0D7DE"),
    bottom=Side(style="thin", color="D0D7DE"),
)


def _aliases(raw) -> str:
    if not raw:
        return ""
    if isinstance(raw, str):
        return raw
    return ", ".join(str(a) for a in raw)


def _example(raw) -> str:
    if raw is None:
        return ""
    return str(raw).strip()


def load_registry(rel: str) -> list[dict]:
    path = REG / rel
    data = yaml.safe_load(path.read_text(encoding="utf-8")) or {}
    rows = []
    for name, spec in data.items():
        spec = spec or {}
        rows.append(
            {
                "name": name,
                "role": spec.get("role") or "",
                "enabled": "yes" if spec.get("enabled", True) else "no",
                "aliases": _aliases(spec.get("aliases")),
                "description": (spec.get("description") or "").strip(),
                "example": _example(spec.get("example")),
                "min_args": spec.get("min_args", ""),
                "requires_value": "yes" if spec.get("requires_value") else "",
                "module": spec.get("module") or "",
                "attr": spec.get("attr") or "",
            }
        )
    return rows


def style_header(ws: Worksheet, ncols: int) -> None:
    fill = PatternFill("solid", fgColor=NAVY)
    font = Font(bold=True, color=WHITE, name="Calibri", size=11)
    for col in range(1, ncols + 1):
        cell = ws.cell(1, col)
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")
        cell.border = THIN
    ws.freeze_panes = "A2"
    ws.row_dimensions[1].height = 22


def style_body(ws: Worksheet, ncols: int) -> None:
    wrap = Alignment(wrap_text=True, vertical="top")
    font = Font(name="Calibri", size=10)
    alt = PatternFill("solid", fgColor=LIGHT)
    for r in range(2, ws.max_row + 1):
        longest = 1
        for c in range(1, ncols + 1):
            cell = ws.cell(r, c)
            cell.alignment = wrap
            cell.font = font
            cell.border = THIN
            if r % 2 == 0:
                cell.fill = alt
            longest = max(longest, str(cell.value or "").count("\n") + 1)
        ws.row_dimensions[r].height = min(140, 16 + 13 * longest)


def autosize(ws: Worksheet, widths: list[int]) -> None:
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


def write_sheet(wb: Workbook, title: str, headers: list[str], rows: list[list], widths: list[int]) -> Worksheet:
    ws = wb.create_sheet(title)
    for i, h in enumerate(headers, start=1):
        ws.cell(1, i, h)
    for ri, row in enumerate(rows, start=2):
        for ci, val in enumerate(row, start=1):
            ws.cell(ri, ci, val)
    style_header(ws, len(headers))
    style_body(ws, len(headers))
    autosize(ws, widths)
    if rows:
        ref = f"A1:{get_column_letter(len(headers))}{1 + len(rows)}"
        safe = "".join(ch if ch.isalnum() else "_" for ch in title)[:30]
        table = Table(displayName=safe, ref=ref)
        table.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True)
        ws.add_table(table)
    return ws


def cover(wb: Workbook) -> None:
    ws = wb.active
    ws.title = "Read me"
    ws["A1"] = "KPI Engine — Capability Catalog"
    ws["A1"].font = Font(name="Calibri", size=22, bold=True, color=NAVY)
    ws["A2"] = (
        "Live allowlist from kpi_engine/registries/. "
        "YAML may only name what is listed here. Generated for authors, architects, and leadership."
    )
    ws["A2"].alignment = Alignment(wrap_text=True)
    ws.merge_cells("A2:F2")
    ws.row_dimensions[2].height = 36

    headers = ["Sheet", "What it lists", "Where you use it in YAML", "How to extend"]
    rows = [
        [
            "Framework components",
            "What ops, hooks, column/measure functions, base_measures, cuts ARE",
            "Read first — pick the right layer before writing YAML",
            "N/A — conceptual",
        ],
        [
            "Naming conventions",
            "File names, kpi_id / model_id, identifiers, aliases, reserved words",
            "Name every YAML file and key before writing formulas",
            "N/A — engine-enforced; do not invent a second spelling",
        ],
        [
            "YAML preparation",
            "AI authoring contract: intake, calculation→YAML map, completeness checklist",
            "Attach this sheet + catalogs + filled intake; AI must emit entire KPI YAML",
            "kpi-yaml-preparation-guide.md §0 — do not invent ops; ask if intake is incomplete",
        ],
        [
            "Decision guide",
            "If you need X, use Y — stop at the first matching row",
            "Walk top to bottom while authoring",
            "If nothing matches: new catalog name, not pipeline/",
        ],
        [
            "Sample dataset",
            "Tiny order fact used by the worked examples (hand-checkable). Anchor = Mar 2026.",
            "N/A — teaching data, not production",
            "N/A",
        ],
        [
            "Sample monthly totals",
            "Oracle rollups of the sample (by month × region, 3m, worldwide)",
            "Use with Worked examples to check arithmetic",
            "N/A",
        ],
        [
            "Worked examples",
            "Same dataset: YAML + arithmetic + expected JSON numbers",
            "Copy YAML into a KPI; numbers are the oracle for March 2026",
            "N/A",
        ],
        [
            "Measure ops",
            "Every measures.op kind (point, window, rank, lag, …)",
            "measures.<key>.op:",
            "capabilities/ops/ + registries/ops.yaml",
        ],
        [
            "Column functions",
            "Row-level functions on retrieved columns",
            "base_measures.*.columns + op:  or  expr: date_diff(…)",
            "capabilities/functions/column/impl.py + registries/functions/column.yaml",
        ],
        [
            "Measure functions",
            "Scalar functions on already-aggregated measures",
            "measures.*.op: fn  /  arithmetic  /  expr:",
            "capabilities/functions/measure/impl.py + registries/functions/measure.yaml",
        ],
        [
            "Hooks",
            "Series stats over the densified period spine",
            "measures.*.op: hook  +  hook: <name>",
            "capabilities/hooks/impl.py + registries/hooks.yaml",
        ],
        [
            "All operations",
            "Aggs, OVER windows, HAVING, filters, time, cuts",
            "base_measures / having / over / time / cuts",
            "Rare: contracts.py + core (new agg, filter op, time format)",
        ],
        [
            "YAML calculation patterns",
            "How to write typical KPI math (YoY, ratio, SLA, rank, …)",
            "Copy the pattern into kpi_config/kpis/<kpi_group>/<id>.yaml",
            "If no pattern fits, add a catalog name — do not edit pipeline/",
        ],
        [
            "How to extend",
            "Step-by-step for each capability type",
            "N/A — engine packaging",
            "Then regenerate registries/CAPABILITIES.md",
        ],
    ]
    for i, h in enumerate(headers, start=1):
        ws.cell(4, i, h)
    for ri, row in enumerate(rows, start=5):
        for ci, val in enumerate(row, start=1):
            ws.cell(ri, ci, val)
    style_header(ws, 4)
    # restyle from row 4
    fill = PatternFill("solid", fgColor=NAVY)
    font = Font(bold=True, color=WHITE, name="Calibri", size=11)
    for col in range(1, 5):
        cell = ws.cell(4, col)
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(wrap_text=True, vertical="center")
        cell.border = THIN
    wrap = Alignment(wrap_text=True, vertical="top")
    alt = PatternFill("solid", fgColor=LIGHT)
    for r in range(5, 5 + len(rows)):
        for c in range(1, 5):
            cell = ws.cell(r, c)
            cell.alignment = wrap
            cell.font = Font(name="Calibri", size=10)
            cell.border = THIN
            if r % 2 == 0:
                cell.fill = alt
        ws.row_dimensions[r].height = 48
    autosize(ws, [28, 52, 48, 62])
    ws.freeze_panes = "A5"

    last = 4 + len(rows)
    ws["A" + str(last + 2)] = (
        "Start here: Framework components → Naming conventions → YAML preparation "
        "→ Decision guide → Sample dataset → Worked examples."
    )
    ws["A" + str(last + 2)].font = Font(bold=True, size=12, color=NAVY)
    ws.merge_cells(f"A{last + 2}:D{last + 2}")

    ws["A" + str(last + 3)] = "Two calculation layers (do not mix them up)"
    ws["A" + str(last + 3)].font = Font(bold=True, size=14, color=NAVY)
    ws.merge_cells(f"A{last + 3}:D{last + 3}")
    ws["A" + str(last + 4)] = (
        "1) base_measures — per retrieved row, then fold with agg:. Identifiers are physical columns "
        "(and earlier helpers). Example: line_ratio: { expr: shipped / ordered, agg: sum } is the SUM of per-row ratios.\n"
        "2) measures — after aggregation, on the monthly/cut row. Identifiers are other measure keys. "
        "Example: fill_rate: { op: expr, expr: shipped_now / ordered_now } is SUM(shipped) / SUM(ordered).\n\n"
        "Prefer columns:+op: when a named function is clearer. Prefer expr: for nested + - * / and CASE. "
        "Prefer op: fn when operand names must not be swapped. Prefer a hook when you need the whole densified series."
    )
    ws["A" + str(last + 4)].alignment = Alignment(wrap_text=True, vertical="top")
    ws.merge_cells(f"A{last + 4}:D{last + 6}")
    ws.row_dimensions[last + 4].height = 80

    ws["A" + str(last + 8)] = (
        "Host entry: kpi_engine.main   |   YAML root: KPI_ENGINE_CONFIG_DIR or sibling kpi_config/   |   "
        "Routing: execution.kpi_id → kpis/<kpi_group>/<id>.yaml   |   "
        "AI brief: attach YAML preparation + catalogs + calculation intake (kpi-yaml-preparation-guide.md §0)."
    )
    ws["A" + str(last + 8)].font = Font(italic=True, size=10, color="44546A")
    ws.merge_cells(f"A{last + 8}:D{last + 8}")


def convention_rows() -> list[list]:
    """File names, identifiers, and other author rules the engine actually enforces."""
    return [
        [
            "KPI YAML file",
            "Path: kpi_config/kpis/<kpi_group>/<kpi_id>.yaml or flat kpis/<kpi_id>.yaml. Stem equals execution.kpi_id exactly (no fold). kpi_id is unique across groups.",
            "kpis/sotif/3004.yaml when execution.kpi_id is 3004. Copy to kpis/<kpi_group>/<new_id>.yaml.",
            "sotif.yaml for kpi_id 3004. 3004.yml. Two groups both containing 3004.yaml. kpis/sotif/quality/3004.yaml (only one group folder).",
            "load_kpi finds exactly one f\"{kpi_id}.yaml\" under kpis/ or kpis/*/. Zero or two files is a bind error. No execution.kpi_group field.",
        ],
        [
            "kpi_group folder",
            "One optional snake_case directory under kpis/ and models/. Authoring only; not sent on the context. Group names use the identifier alphabet.",
            "kpis/sotif/, models/sotif/, models/freight/. A KPI in kpis/sotif/ may model: a file in models/freight/.",
            "execution.kpi_group. Nested groups (kpis/a/b/id.yaml). Spaces or hyphens in the folder name.",
            "The engine never reads the folder name. Duplicate kpi_id or model_id across groups is a bind error.",
        ],
        [
            "KPI yaml kpi_id:",
            "Set kpi_id: to the same value as the filename stem (and the host execution.kpi_id).",
            "File 3004.yaml → kpi_id: 3004. File freight_fill.yaml → kpi_id: freight_fill.",
            "File 3004.yaml with kpi_id: sotif. Reusing one YAML for several ids.",
            "The engine loads by filename. Inner kpi_id: is the spec id; keep it identical so logs and errors match the request.",
        ],
        [
            "Model YAML file",
            "Path: kpi_config/models/<kpi_group>/<model_id>.yaml or flat models/<model_id>.yaml. Prefer lowercase snake_case. Extension must be .yaml. model_id is unique across groups.",
            "models/sotif/sotif.yaml, models/freight/orders_sql.yaml.",
            "sotif.yml. models/Sotif Model.yaml. Two groups both containing sotif.yaml. Putting the file under kpis/.",
            "load_model finds exactly one file whose stem folds to model_id among models/*.yaml and models/*/*.yaml (Sotif.yaml ↔ sotif). Two fold-matches is an error.",
        ],
        [
            "Model yaml model_id:",
            "model_id: must equal the filename stem you intend KPIs to point at. Prefer the same spelling everywhere.",
            "File sotif.yaml → model_id: sotif. File orders_sql.yaml → model_id: orders_sql.",
            "File sotif.yaml with model_id: sotif_sql. Two files that fold to the same id.",
            "KPI model: is matched to this field after folding case / spaces / underscores. Extra tokens do not fold away: sotif ≠ sotif_sql.",
        ],
        [
            "KPI model: pointer",
            "KPI YAML model: must name a model_id that exists. Same fold as model_id (case, spaces, underscores only).",
            "model: sotif  with models/sotif/sotif.yaml (or models/freight/sotif.yaml). Per-base override base_measures.x.model: orders_sql when two extracts are needed.",
            "model: sotif pointing at sotif_sql.yaml. Embedding an ADLS path instead of a model id. Inventing a model that has no file.",
            "same_model_id folds Region/region and my_model/my model. It does not treat sotif as sotif_sql. Physical vs sql is kind: on the model file, not a suffix rule — but keep suffixes honest (orders_sql.yaml for kind: sql).",
        ],
        [
            "Identifier alphabet",
            "YAML names that become SQL/Pandas identifiers must match ^[A-Za-z_][A-Za-z0-9_]*$. Prefer snake_case.",
            "reason_code, current_value, event_month, sotif_value, shipped_now.",
            "reason-code, reason.code, \"Reason Code\", 12m_trend, current value.",
            "require_ident rejects hyphens, dots, spaces, quotes, and leading digits. Used for dimensions, base_measures, parameters, aliases, sources, join keys, time.column, filter_map values, output_schema, columns: lists.",
        ],
        [
            "Reserved words in formulas",
            "Do not use these as column or measure names inside expr: / CASE: case, when, then, else, and, or, not, is, null, in.",
            "date_diff(start, end, 'day') — end is legal. Names like region, amount, shipped_now.",
            "expr: case / qty. A column named when or in.",
            "Parser treats those as CASE/boolean keywords. end is a CASE terminator but is allowed as an identifier (date_diff(start, end, 'day')).",
        ],
        [
            "Name folding (host ↔ YAML)",
            "Host keys fold onto YAML keys: lowercase, trim, spaces → underscore. Measure keys also try a compact fold (drop remaining underscores).",
            "YAML current_value matches host Current_Value or current value. YAML previous_year_value matches PreviousYearValue.",
            "Relying on fold for file names (KPI files do not fold). Expecting sotif to match sotif_sql. Two YAML keys that fold to the same spelling.",
            "norm_name is case/space/underscore only. Use one canonical snake_case in YAML; the host may send UI Title Case. Do not create two keys that collide after fold.",
        ],
        [
            "Measure keys",
            "Keys under measures: are what context.measures_required[].measure_key must match (after fold). snake_case. Must not collide with parameters.",
            "current_value, previous_year_value, trailing_3m, yoy_pct, fill_rate.",
            "Current Value as the YAML key. Reusing a parameter name. A key that is not in the host contract.",
            "Unknown measure_key is a bind error listing valid YAML keys. An empty measures_required list computes nothing (it does not expand to every key).",
        ],
        [
            "Base measure names",
            "Keys under base_measures: are internal facts. UI does not request these names. snake_case identifiers.",
            "sotif_value, shipped_qty, ordered_qty, line_rev.",
            "Naming a base the same as a requestable measure unless you intend identity. Hyphenated names.",
            "Measures reference bases via of:. A base with no agg: is a row helper, not a requestable point.",
        ],
        [
            "Dimensions",
            "Catalog names under dimensions[].name (and from: if the physical column differs). Request selected_dimensions must pick from this allowlist.",
            "- { name: supplier, from: supplier_name }. default_dimensions: [reason_code].",
            "Inventing a request dimension not in YAML. Using the host display label as the YAML name when the column is supplier_name (map it with from:).",
            "YAML owns the allowlist. from: is the physical/model column; name: is the grain token the request uses.",
        ],
        [
            "Cut names",
            "Short stable tokens. G and R in 3004 are examples, not engine-hardcoded. default_cut must be one of the declared names.",
            "cuts: [{ name: G, … }, { name: R, group_by: [region] }]. default_cut: G.",
            "Renaming G/R in YAML without updating the host cut list. Two cuts with the same name.",
            "Cut name is a string the host sends; group_by entries must be catalog dimensions (identifiers). Also_emit / ignore_filters name other cuts or filter codes.",
        ],
        [
            "Dataset aliases",
            "model.required_aliases and sources.*.alias must be identifiers. They bind to context.datasets by alias or key, case-insensitive.",
            "required_aliases: [sotif]. Host dataset alias or key: sotif (or SOTIF).",
            "required_aliases: [my-table]. Putting an ADLS URI in the alias. A KPI-level path instead of an alias.",
            "Context path wins; model default_path / default_paths fills gaps. Do not put dataset URIs in KPI YAML.",
        ],
        [
            "SQL model placeholders",
            "kind: sql models reference bound datasets as $alias_path or $alias_scan (alias = required_aliases entry).",
            "sql: |\n  SELECT … FROM $sotif_scan WHERE …",
            "Hard-coding /mnt/… or abfss:// in the SQL. Using an alias that is not in required_aliases.",
            "$alias_scan becomes delta_scan(?) or read_parquet(?) from table_type. $alias_path is the path parameter. Order of placeholders must match bind order.",
        ],
        [
            "Time grains",
            "Only: day, week, month, quarter, year (lowercase). Default grain is month.",
            "time: { column: event_month, grain: month, filter_code: reporting_month }.",
            "time.grain: monthly. MTD as a grain name. execution.time_grain on the host payload.",
            "Send parameters.time_grain when the KPI declares that parameter. execution.time_grain is rejected. calendar: gregorian or fiscal.",
        ],
        [
            "time.filter_code and filter_map keys",
            "These are host filter codes, not SQL identifiers. They may contain spaces. filter_map values (columns) must still be identifiers.",
            "filter_code: reporting_month. periods: { year: year, month: \"current month\" }. filter_map: { \"Supplier Name\": supplier_name }.",
            "filter_map values with spaces or hyphens. Setting periods: and compose: together.",
            "time.filter_code is required unless periods: or compose: is set. A scalar filter_code on the context still wins (one period, not a month IN list). periods: parts are independent predicates.",
        ],
        [
            "Parameters",
            "Parameter names are identifiers. Must not equal a measure key. Do not declare selected_dimensions as a parameter. when: case labels cannot be param, cases, or else.",
            "parameters: { time_grain: { type: string, allowed: [month, quarter] } }.",
            "parameters.selected_dimensions. A case label named else. A parameter named current_value when that is also a measure key.",
            "time_grain is the reserved overlay formerly on execution. Case labels param / cases / else are when: metadata keys.",
        ],
        [
            "Host UDF / module_path",
            "One generic entry: kpi_engine.main. Routing is execution.kpi_id → kpi_config/kpis/<kpi_group>/<id>.yaml (or flat kpis/<id>.yaml). Do not add a per-KPI Python module.",
            "Host module_path: kpi_engine.main. New KPI = new YAML (+ model YAML if the extract is new).",
            "kpi_engine/<kpi_id>/main.py. Cloning the engine per metric. A second module_path per kpi_id.",
            "The UDF is stateless compute(context). KPI math lives in YAML; DuckDB shape lives in model YAML.",
        ],
        [
            "What must not live in KPI YAML",
            "No dataset URIs, no Python, no invented op:/hook:/fn: names, no per-cut model:.",
            "model: sotif. op: point / window / rank / hook / expr — names from this workbook's catalogs.",
            "abfss://… in the KPI file. op: my_custom_yoy. cuts: [{ model: other }]. A sibling .py next to the YAML.",
            "Paths belong on context.datasets or model default_paths. Catalog names are the Measure ops / Functions / Hooks sheets. cuts cannot set model:.",
        ],
        [
            "File header comments (recommended)",
            "Match existing config files: what the file provides, where it is used, capabilities, when to copy it.",
            "See kpi_config/kpis/sotif/3004.yaml and kpi_config/models/sotif/sotif.yaml headers.",
            "Leaving a copied file with the old kpi_id / Sotif description. Putting secrets in comments.",
            "Comments are documentation only; the engine ignores them. Keep kpi_id / model_id in the body in sync with the filename.",
        ],
    ]


def preparation_rows() -> list[list]:
    """How to brief an AI so it emits a complete, bind-ready KPI YAML."""
    ops = ", ".join(r["name"] for r in load_registry("ops.yaml"))
    col_fns = ", ".join(r["name"] for r in load_registry("functions/column.yaml"))
    meas_fns = ", ".join(r["name"] for r in load_registry("functions/measure.yaml"))
    hooks = ", ".join(r["name"] for r in load_registry("hooks.yaml"))
    return [
        [
            "How to brief an AI",
            "Attach this workbook (Naming conventions, this sheet, Measure ops, Column/Measure functions, Hooks, YAML patterns) plus kpi-yaml-preparation-guide.md §0 plus a filled intake.",
            "One message: catalogs + intake + 'emit complete YAML, no TODOs'.",
            "A one-line 'write YAML for fill rate' with no kpi_id, columns, or measure_keys.",
            "The AI may only use names on the catalog sheets. If intake is incomplete it must ask, not guess.",
        ],
        [
            "AI output contract",
            "Every generating response is: (1) entire kpis/<kpi_group>/<kpi_id>.yaml (2) entire models/<kpi_group>/<model_id>.yaml if extract is new, else 'reuses existing model' (3) Assumptions (4) Gaps.",
            "Full files with header comments updated for this kpi_id. Then a short assumptions/gaps list.",
            "A snippet, skeleton with blanks, # TODO, CHANGE_ME, or '...'. Partial measures: block.",
            "High confidence = completeness checklist all true. If any required key is unknown, ask numbered questions and emit no YAML.",
        ],
        [
            "Intake — identity",
            "kpi_id (filename stem = execution.kpi_id exactly). model_id. reuse_existing_model yes/no. dataset_aliases.",
            "kpi_id: 4120. model_id: sotif (reuse) or freight_lane (new). aliases: [sotif].",
            "Omitting kpi_id. Reusing sotif when the tables are different. Inventing a model_id that folds onto another file.",
            "See Naming conventions. .yaml only. Host module_path stays kpi_engine.main.",
        ],
        [
            "Intake — time or snapshot",
            "Either time.column + grain (day|week|month|quarter|year) plus filter_code and/or periods:, or snapshot: true (omit time:).",
            "time: { column: event_month, grain: month, periods: { year: year, month: \"current month\" }, calendar: gregorian }.",
            "Guessing reporting_month. execution.time_grain. Setting periods: and compose: together. A part finer than time.grain.",
            "Scalar filter_code on the context is one value and wins. periods: parts conjoin (year alone = full year). Missing parts probe history. Snapshot forbids window/trend/nonzero offset.",
        ],
        [
            "Intake — grain and cuts",
            "dimensions (name + from physical column). default_dimensions (required, [] = worldwide). At least one cut. default_cut. output_cut is a walk root (pack_also_emit: false locks one cut).",
            "dimensions: [{ name: region, from: region }]. cuts G extras-only group_by; exclude_from_grain / ignore_filters as needed.",
            "Inventing a host dimension not in YAML. Putting default_dimensions names again in cuts.group_by. Treating G/R as engine-hardcoded.",
            "YAML owns the allowlist. Request selected_dimensions only picks from it. Do not declare parameters.selected_dimensions.",
        ],
        [
            "Intake — calculations",
            "Every host measure_key plus English/math. Physical columns DuckDB must retrieve. Optional filters: declarations.",
            "measure_keys: [current_value, yoy_pct]. calculations.current_value: SUM(amount) at selected month.",
            "A calculation with no measure_key. A measure_key with no formula. Guessing column amount vs billed.",
            "Unknown measure_key is a bind error. Empty measures_required computes nothing. Helpers the UI does not request go under base_measures only.",
        ],
        [
            "Map — current period total",
            "SUM/COUNT/MIN/MAX of a column over the time selection (one bucket or many).",
            "base_measures.fact: { sql: amount, agg: sum }\nmeasures.current_value: { of: fact, op: point, offset: { months: 0 } }",
            "A measure op for the raw column. Free SQL SUM() in the KPI file.",
            "agg: sum | avg | count | count_distinct | min | max | median | percentile | first | last.",
        ],
        [
            "Map — previous period / YoY",
            "Same metric last year/month/quarter over the shifted selection; growth % vs that set (year-only is year vs year).",
            "previous_year_value: { of: fact, op: point, offset: { years: 1 } }\nyoy_pct: { op: fn, fn: growth_pct, inputs: [current_value, previous_year_value] }",
            "A window of length 1. Hand-coded (current-previous)/previous in SQL. fn: divide when the spec asked for %.",
            "growth_pct nulls when the base is null or zero. offset units: days weeks months quarters years.",
        ],
        [
            "Map — last period of a composite",
            "Lag/diff/index of a ratio, OEE, or rolling window — not a duplicated prior_* subgraph.",
            "prior_oee: { op: lag, of: oee_pct, offset: { months: 1 } }\ndelta: { op: diff, of: oee_pct, offset: { months: 1 } }",
            "lag of trend/hook/rank. lag of a row helper (no agg:).",
            "Sources: bases and shiftable measures (point, window, fn, expr, …).",
        ],
        [
            "Map — row mask (where:)",
            "COUNT/SUM only the rows that pass a comparison (status, mrr > 0, amount between).",
            "where: { column: mrr, op: gt, value: 0 }\nbetween: { column: amount, op: between, values: [lo, hi] }",
            "like / is_null on base where:. between with only value:.",
            "Ops: in eq ne gt gte lt lte between. ne excludes nulls (SQL-style). Numeric ops coerce the column.",
        ],
        [
            "Map — trailing N / YTD / QTD",
            "Trailing N periods as one number; or named period-to-date.",
            "value_3m: { of: fact, op: window, trailing: { months: 3 }, inclusive: true }\nvalue_qtd: { of: fact, op: window, range: qtd }",
            "Summing three point keys. range: qtd to mean trailing 3. wtd at month grain.",
            "range: trailing | leading | cumulative | ytd | mtd | qtd | wtd | full_month | full_quarter | full_year. wtd needs day grain.",
        ],
        [
            "Map — graph / trend",
            "Last N periods as an array for a chart.",
            "trend_12m: { of: fact, op: trend, trailing: { months: 12 }, cuts: [G] }",
            "Returning N point keys. Unrestricted trend on a high-cardinality cut (50k cell cap).",
            "Trend defaults to default_cut if cuts: omitted. Axis lands in response trend_axes / trend_labels.",
        ],
        [
            "Map — ratio of totals vs sum of ratios",
            "Fill rate / rate KPIs: decide per-row then SUM vs SUM/SUM.",
            "Two summed bases → two points → measures.fill_rate: { op: expr, expr: shipped_now / ordered_now }",
            "base expr shipped/ordered + agg: sum when the spec wanted ratio of totals.",
            "base expr identifiers are physical columns. measure expr identifiers are other measure keys. Do not mix expr: with columns:/op:/sql: on one base.",
        ],
        [
            "Map — share / rank / SLA",
            "Share of all groups on this cut; rank; hit-rate over a series.",
            "op: percent_of_total (not fn: percent). op: rank. op: hook + hook: hit_rate + trailing: + value:.",
            "fn: percent for cut share. Sorting in the host to fake rank. A window SUM for 'how often'.",
            "Cut ops see the full cut. Hooks need the densified spine — declare trailing so required_span is wide enough.",
        ],
        [
            "Map — snapshot / entity window / lookup",
            "No period column; per-entity sequence; static code map.",
            "Omit time:. over: on a pre-fold base (not calendar op: lag). lookup: { column, map, default }.",
            "window/trend on a snapshot. Calendar op: lag for per-customer previous order. Giant CASE for a fee table.",
            "Snapshot leftover reporting_month skips as no_time. identity_grain required to point at a helper with no agg:.",
        ],
        [
            "Map — cannot express",
            "Iterative allocation, custom solver, anything with no catalog row.",
            "Stop. Report: cannot bind with current catalog. Name the missing op/fn/hook.",
            "Inventing op: my_yoy. eval() in YAML. A per-KPI Python module. Editing pipeline/.",
            "New reusable names go in capabilities/ + registries/ (How to extend sheet). Do not fake them in this KPI file.",
        ],
        [
            "Mandatory KPI keys",
            "kpi_id, model, default_dimensions, ≥1 cut, ≥1 measure. time: (period) or omit entirely (snapshot). row_set span_union|anchor_only.",
            "See gold file kpi_config/kpis/sotif/3004.yaml — copy structure, not Sotif names, unless this KPI is Sotif.",
            "Missing default_dimensions. Empty measures:. cuts[].model:. parameters.selected_dimensions.",
            "cuts.group_by is extras only. Header comments must name this kpi_id, not leftover 3004 text.",
        ],
        [
            "Mandatory model keys (new extract only)",
            "model_id, kind physical|sql, required_aliases, sources (physical) or sql:+output_schema (sql).",
            "model_id matches filename and KPI model:. SQL uses $alias_scan / $alias_path, not hardcoded URIs.",
            "KPI-level sql: as the extract. A second model whose stem folds onto an existing file.",
            "Context path wins over default_paths. Prefer $alias_scan so Delta vs Parquet follows table_type.",
        ],
        [
            "Completeness checklist",
            "Emit YAML only when every box is true (same list as kpi-yaml-preparation-guide.md §0.7).",
            "Filename=kpi_id=execution.kpi_id; model matches; every host measure_key declared; every of:/expr ident exists; identifiers legal; time xor snapshot rules; no URIs; no invented names.",
            "Shipping 'best effort' YAML with guessed columns. Leaving one measure_key unimplemented.",
            "If any box is false: numbered questions, no files. Closed-world names below (catalog sheets win if they disagree).",
        ],
        [
            "Closed-world — measure ops",
            "measures.op may only be one of these names (live from registries/ops.yaml).",
            ops,
            "Any other op: string, including aliases you invent.",
            "See Measure ops sheet for YAML shape. Aliases listed there (if any) are the only extra spellings.",
        ],
        [
            "Closed-world — column functions",
            "base_measures columns:+op: / expr call names (live from registries/functions/column.yaml).",
            col_fns,
            "SQL functions that are not on this list.",
            "See Column functions sheet.",
        ],
        [
            "Closed-world — measure functions",
            "op: fn / arithmetic fn: names (live from registries/functions/measure.yaml).",
            meas_fns,
            "fn: yoy as a new name — use growth_pct (alias yoy/mom if listed on that sheet).",
            "See Measure functions sheet.",
        ],
        [
            "Closed-world — hooks",
            "op: hook + hook: names (live from registries/hooks.yaml). Always declare trailing/offset so the scan is wide enough.",
            hooks,
            "A hook name that is not listed. Using a window SUM as a substitute for hit_rate.",
            "See Hooks sheet. Hooks that list requires_value need value:.",
        ],
    ]


def component_rows() -> list[list]:
    return [
        [
            "KPI YAML",
            "The definition file for one kpi_id",
            "Declares time, dimensions, base_measures, cuts, measures. This is what authors write.",
            "kpi_config/kpis/<kpi_group>/<kpi_id>.yaml (or flat kpis/<kpi_id>.yaml)",
            "Not Python. Not a per-KPI UDF.",
            "When you have a new metric the UI will request.",
        ],
        [
            "Model YAML",
            "What DuckDB reads (tables/joins or a SQL CTE)",
            "Shapes the extract. Does not own KPI math (YoY, rank, windows).",
            "kpi_config/models/<kpi_group>/<id>.yaml  —  KPI model: pointer",
            "Do not put YoY in model SQL as the default path.",
            "New source tables, eligibility CTE, timezone conversion.",
        ],
        [
            "base_measures",
            "Internal facts computed per retrieved row, then folded",
            "sql: / expr: / lookup: / over: / columns+op. Identifiers are physical columns (or earlier helpers). agg: folds rows (sum, avg, …).",
            "base_measures:\n  gross: { expr: qty * price }\n  billed: { expr: gross, agg: sum }",
            "The UI usually does not request these names. measures: of: them.",
            "Need qty×price, a fee lookup, or a per-customer running sum before GROUP BY.",
        ],
        [
            "Column function",
            "A named row-wise function on retrieved columns (pandas Series)",
            "Runs once per fact row. Examples: multiply, divide, coalesce, date_diff. Registered in functions/column.yaml.",
            "base_measures:\n  line_rev:\n    columns: [qty, price]\n    op: multiply\n    agg: sum",
            "Not a total-of-totals. SUM(qty*price) ≠ SUM(qty)*SUM(price).",
            "Named math on columns is clearer than a formula, or you will reuse the function.",
        ],
        [
            "Measure (output)",
            "A requestable column in the JSON (measure_key)",
            "Every UI key must exist under measures:. After aggregation, on the monthly/cut row.",
            "measures:\n  current_value:\n    of: billed\n    op: point",
            "Empty measures_required computes nothing.",
            "The host asks for this name.",
        ],
        [
            "Measure op (kind)",
            "The shape of a measure: point, window, rank, hook, …",
            "Tells the engine HOW to produce the value from of:/inputs:. Registry: ops.yaml. Combo ops run per group; cut ops see all groups; period ops need time:.",
            "op: point | window | trend | fn | expr | rank | hook | …",
            "Do not invent op names. Unknown op is BindError.",
            "See Decision guide: pick the kind first, then fill fields.",
        ],
        [
            "Measure function",
            "A named scalar function on already-aggregated measure values",
            "Inputs are other measure keys (totals), not physical columns. Registry: functions/measure.yaml. Called via op: fn, arithmetic, or expr.",
            "yoy:\n  op: fn\n  fn: growth_pct\n  inputs: [current_value, previous_year_value]",
            "growth_pct(40, 50) = (40-50)/50 = -0.20. Null/zero base → null.",
            "YoY, ratio of two totals, attainment, clamp, sign_label.",
        ],
        [
            "Hook",
            "A named function over the densified period series for one group",
            "Needs many months of history (trailing/offset). Examples: ewma, hit_rate, cagr, slope. Registry: hooks.yaml. Does not open DuckDB.",
            "smoothed:\n  op: hook\n  hook: ewma\n  of: billed\n  trailing: { months: 6 }",
            "Not the same as op: window (one number from SUM). A hook can use every month in the window.",
            "Hit-rate, streaks, EWMA, CAGR, volatility — stats on the time series.",
        ],
        [
            "Cut",
            "A named grouping grain + which filters to ignore",
            "G/R are examples. group_by, ignore_filters, also_emit, pack_also_emit. Request grain overlays selected_dimensions. Context output_cut is the walk root.",
            "cuts:\n  - { name: G, group_by: [], ignore_filters: [region], also_emit: [R] }\n  - { name: R, group_by: [region] }",
            "Global avg is recomputed, never mean of regional avgs. YAML default for output_cut does not lock.",
            "Worldwide vs by-region in one response.",
        ],
        [
            "Calendar lag vs entity over",
            "Two different 'lag' concepts",
            "op: lag = same measure on the densified month spine vs the anchor. over.fn: lag = previous ORDER row for a customer.",
            "Calendar: of: value_3m, op: lag, offset: { years: 1 }\nEntity: over: { fn: lag, of: order_date, partition_by: [customer_id], order_by: [order_date] }",
            "Mixing them is the usual bug.",
            "Last year's 3m window vs days since last order.",
        ],
        [
            "having vs predicate",
            "Drop groups vs flag groups",
            "having: removes rows from JSON. op: predicate: keeps rows with 1/0.",
            "having: { predicates: [{ of: current_value, cmp: gt, value: 0 }] }\nhealthy: { op: predicate, predicates: [...] }",
            "Densified zeros can fail having gt: 0.",
            "Drop empty months vs show a health flag.",
        ],
        [
            "identity_grain",
            "When a row helper can be measures.of",
            "Every emitted cut's grain must equal identity_grain. Duplicate tuples → CatalogError.",
            "identity_grain: [order_id]\nmeasures:\n  net_now: { of: net_charge, op: point }",
            "Coarser also_emit cut → BindError. Fold with agg: instead.",
            "Per-order SLA fields without agg: max workarounds.",
        ],
        [
            "Capability / registry",
            "The allowlist of names YAML may call",
            "A name is callable only if listed and enabled. Live page: CAPABILITIES.md.",
            "registries/ops.yaml, functions/*.yaml, hooks.yaml",
            "Dotted import paths and context.udf.module_path are rejected.",
            "Extending the engine without editing pipeline/.",
        ],
        [
            "pipeline/ (frozen pipeline)",
            "adapt → bind → extract → calculate → JSON",
            "Not for a new catalog name. Change only for new agg, filter operator, time format, or common YAML field.",
            "kpi_engine/pipeline/",
            "No if kpi_id == … branches.",
            "Rare platform work, not KPI onboarding.",
        ],
    ]


def decision_rows() -> list[list]:
    return [
        [
            "1",
            "I need to join tables, filter eligibility, or reshape messy source SQL",
            "Model YAML (kind: physical or kind: sql)",
            "KPI sql: as free SQL / SUM() in KPI formulas",
            "Model owns retrieve. KPI YAML owns measure math after retrieve.",
        ],
        [
            "2",
            "I need one physical column, then SUM/AVG/COUNT the rows",
            "base_measures + sql: + agg:",
            "A measure op: for the raw fact",
            "sotif_value: { sql: amount, agg: sum }",
        ],
        [
            "3",
            "I need qty × price (or similar) on each row, then SUM",
            "Column function (columns+op: multiply) OR base expr:, then agg: sum",
            "SUM(qty) * SUM(price) as a measure expr",
            "See sample: line_revenue = qty*price; billed = sum(line_revenue).",
        ],
        [
            "4",
            "I need nested + - * / or CASE on columns",
            "base_measures named expr: steps",
            "SQL CASE in DuckDB inside KPI sql:",
            "CASE WHEN discounted < 150 THEN shipping ELSE 0 END",
        ],
        [
            "5",
            "I need a static code→value map (COD fee, tier rebate)",
            "lookup: on a base",
            "Giant CASE for 3 keys",
            "lookup: { column: pay, map: { COD: 25 }, default: 10 }",
        ],
        [
            "6",
            "I need per-customer sequence, days since last order, running billed",
            "over: on pre-fold detail (entity window)",
            "Calendar op: lag (that is vs the month spine / anchor)",
            "partition_by + order_by; of: required for lag/running_*/last_n",
        ],
        [
            "7",
            "I need one number for the selected month (or last year / last quarter)",
            "op: point + offset",
            "A window of length 1 unless you want window null/zero rules",
            "offset: { years: 1 } is previous year, not 12 rows back",
        ],
        [
            "8",
            "I need trailing 3/6/12 periods or QTD/YTD as ONE number",
            "op: window (trailing: or range: qtd/ytd/mtd)",
            "Adding several point measures by hand",
            "range: qtd is quarter-to-date, not trailing 3 months",
        ],
        [
            "9",
            "I need a graph (array of months)",
            "op: trend",
            "Many point keys",
            "Cap 50,000 cells/cut; default_cut only unless cuts: listed",
        ],
        [
            "10",
            "I need YoY % or ratio of two already-computed measures",
            "Measure function: op: fn (growth_pct, divide, attainment) or op: expr",
            "Repeating the formula on base_measures",
            "Inputs are measure keys, not column names.",
        ],
        [
            "11",
            "I need this group's share of ALL groups on the cut",
            "op: percent_of_total (optional partition_by)",
            "fn: percent (that only sees this row's two inputs)",
            "partition_by ⊆ cut grain",
        ],
        [
            "12",
            "I need rank / top 3 / quartiles / vs leader",
            "Cut ops: rank, top_n, ntile, gap_to_leader, …",
            "Sorting in the host after the fact if the engine must rank",
            "Ties: rank skips; dense_rank does not",
        ],
        [
            "13",
            "I need last year's trailing-3m (same window, prior year)",
            "op: lag of the window measure, offset years: 1",
            "Entity over.lag",
            "Calendar lag vs densified spine",
        ],
        [
            "14",
            "I need hit-rate, streak, EWMA, CAGR, volatility on monthly values",
            "Hook (op: hook + trailing so the scan is wide enough)",
            "Editing pipeline/ per KPI",
            "requires_value hooks need value: (e.g. hit_rate value: 95)",
        ],
        [
            "15",
            "I need to DROP groups below a floor",
            "KPI having:",
            "op: predicate (that only flags) or apply: result dim-only",
            "Densified 0 can fail gt: 0 — often what you want",
        ],
        [
            "16",
            "I need a 1/0 health flag but keep the row",
            "op: predicate",
            "having: (that deletes the row)",
            "match: all | any",
        ],
        [
            "17",
            "I need a fixed target / goal number",
            "op: constant",
            "Hard-coding the number in every expr",
            "Then vs_target or fn: attainment",
        ],
        [
            "18",
            "I need Positive / Negative / Neutral",
            "fn: sign_label",
            "green_when (that is a boolean bar, not a label)",
            "Zero → Neutral; null stays null",
        ],
        [
            "19",
            "I need per-order fields as JSON columns (no SUM)",
            "identity_grain + point of a helper",
            "agg: max as a workaround, or a coarser cut",
            "Every emitted cut must equal identity_grain",
        ],
        [
            "20",
            "I need math no listed name can express (iterative, ML, geo)",
            "New hook or a kind: sql model",
            "eval(), importlib paths, per-KPI Python",
            "capabilities/ + registries/ then regen CAPABILITIES.md",
        ],
    ]


def sample_fact_rows() -> list[list]:
    """Hand-checkable orders. Anchor = March 2026. Amount = qty * price."""
    return [
        ["2025-03-01", "NA", "LY1", 10, 5, 10, 50, "Last-year March (YoY base for NA)"],
        ["2025-03-01", "EU", "LY2", 5, 8, 5, 40, "Last-year March (YoY base for EU)"],
        ["2026-01-01", "NA", "O1", 10, 5, 10, 50, "Jan NA"],
        ["2026-01-01", "EU", "O2", 4, 8, 3, 32, "Jan EU; shipped < qty"],
        ["2026-02-01", "NA", "O3", 6, 5, 6, 30, "Feb NA"],
        ["2026-02-01", "EU", "O4", 2, 10, 2, 20, "Feb EU"],
        ["2026-03-01", "NA", "O5", 8, 5, 7, 40, "Mar NA; fill 7/8"],
        ["2026-03-01", "EU", "O6", 5, 8, 5, 40, "Mar EU; fill 5/5"],
    ]


def sample_monthly_rows() -> list[list]:
    return [
        ["2025-03", "NA", 50, "LY1 only"],
        ["2025-03", "EU", 40, "LY2 only"],
        ["2026-01", "NA", 50, "O1"],
        ["2026-01", "EU", 32, "O2"],
        ["2026-02", "NA", 30, "O3"],
        ["2026-02", "EU", 20, "O4"],
        ["2026-03", "NA", 40, "O5"],
        ["2026-03", "EU", 40, "O6"],
        ["2026-03", "Worldwide", 80, "NA+EU at anchor"],
        ["2026-01..03", "NA 3m", 120, "50+30+40"],
        ["2026-01..03", "EU 3m", 92, "32+20+40"],
        ["2026-01..03", "Worldwide 3m", 212, "120+92"],
    ]


def worked_example_rows() -> list[list]:
    return [
        [
            "1. Current month revenue (point + column fn)",
            "Column function, then measure op: point",
            "Need SUM(qty*price) at March, by region and worldwide",
            "base_measures:\n  line_rev:\n    columns: [qty, price]\n    op: multiply\n    agg: sum\nmeasures:\n  current_value:\n    of: line_rev\n    op: point\n    offset: { months: 0 }",
            "NA: 8×5=40. EU: 5×8=40. Worldwide: 80.",
            "Do not compute  (8+5)×(5+8). That is SUM(qty)*SUM(price).",
        ],
        [
            "2. Year-over-year % (measure function)",
            "Measure op: point twice, then measure function growth_pct",
            "Need (this March − last March) / last March per region",
            "measures:\n  current_value: { of: line_rev, op: point, offset: { months: 0 } }\n  previous_year_value: { of: line_rev, op: point, offset: { years: 1 } }\n  yoy:\n    op: fn\n    fn: growth_pct\n    inputs: [current_value, previous_year_value]",
            "NA: (40−50)/50 = −0.20 (−20%). EU: (40−40)/40 = 0. Worldwide: (80−90)/90 ≈ −0.111.",
            "growth_pct is a measure function: it sees two scalars, not the fact table. Zero/null last year → null.",
        ],
        [
            "3. Trailing 3-month sum (window op)",
            "Measure op: window",
            "Need Jan+Feb+Mar billed as one number (inclusive of anchor)",
            "measures:\n  value_3m:\n    of: line_rev\n    op: window\n    trailing: { months: 3 }\n    inclusive: true",
            "NA: 50+30+40 = 120. EU: 32+20+40 = 92. Worldwide: 212.",
            "Window is a measure op on the densified month spine, not three point keys added in YAML.",
        ],
        [
            "4. Fill rate of totals (measure expr)",
            "Measure op: expr on two points",
            "Need SUM(shipped)/SUM(qty) in March — ratio of totals",
            "base_measures:\n  shipped_qty: { sql: shipped, agg: sum }\n  ordered_qty: { sql: qty, agg: sum }\nmeasures:\n  shipped_now: { of: shipped_qty, op: point }\n  ordered_now: { of: ordered_qty, op: point }\n  fill_rate:\n    op: expr\n    expr: shipped_now / ordered_now",
            "NA: 7/8 = 0.875. EU: 5/5 = 1.0. Worldwide: 12/13 ≈ 0.923.",
            "If you used base expr shipped/qty then agg: sum you would SUM per-row ratios (NA: 7/8 only one row anyway). Different when many lines.",
        ],
        [
            "5. Share of worldwide (cut op)",
            "Measure op: percent_of_total",
            "Need each region's March revenue as % of all regions",
            "measures:\n  group_share:\n    op: percent_of_total\n    of: current_value",
            "NA: 40/80×100 = 50. EU: 50. (fn: percent would need you to pass part and whole as two inputs on the same row — it cannot see other groups.)",
            "percent_of_total is a cut op: it looks across the cut. Measure function percent only sees its two inputs.",
        ],
        [
            "6. Rank regions (cut op)",
            "Measure op: rank",
            "Need 1 = highest March revenue. Ties share a rank and skip the next.",
            "measures:\n  rev_rank:\n    op: rank\n    of: current_value\n    order: desc",
            "NA 40 and EU 40 → both rank 1; a third group would be 3 (rank skips). dense_rank would give the next 2.",
            "Rank is a cut op. It is not over.rank (entity window on order rows).",
        ],
        [
            "7. Hit SLA 3 of last 6 months (hook)",
            "Hook hit_rate on monthly billed vs 35",
            "Need % of observed months with billed ≥ 35 (teaching threshold)",
            "measures:\n  months_ok:\n    op: hook\n    hook: hit_rate\n    of: line_rev\n    trailing: { months: 3 }\n    value: 35",
            "NA months 50, 30, 40 → 2 of 3 ≥ 35 → 66.67%. EU 32, 20, 40 → 1 of 3 → 33.33%.",
            "A window SUM cannot answer 'how often'. That needs the series → hook. Declare trailing so required_span covers Jan–Mar.",
        ],
        [
            "8. Drop empty densified groups (having)",
            "having, not a function",
            "A region with only February data densifies to 0 in March; drop it",
            "having:\n  predicates:\n    - { of: current_value, cmp: gt, value: 0 }",
            "On this sample both regions have March rows, so both survive. If EU had no March order, EU current_value would be 0 and drop.",
            "op: predicate would keep the row with healthy=0 instead of deleting it.",
        ],
        [
            "9. Per-order fill (identity + column fn)",
            "Column function + identity_grain (no agg)",
            "Need each order_id's shipped/qty as a JSON column",
            "identity_grain: [order_id]\nbase_measures:\n  fill_row:\n    columns: [shipped, qty]\n    op: divide\nmeasures:\n  fill_now: { of: fill_row, op: point }",
            "O5: 7/8=0.875. O6: 5/5=1. O2: 3/4=0.75. Request grain must be order_id (not region).",
            "Without identity_grain, of: fill_row is BindError (row helper). Do not agg: max unless you truly want a fold.",
        ],
        [
            "10. Target attainment (constant + measure fn)",
            "Measure op: constant + measure function attainment",
            "Need actual/target×100 vs a goal of 45 worldwide",
            "measures:\n  target: { op: constant, value: 45 }\n  vs_goal:\n    op: fn\n    fn: attainment\n    inputs: [current_value, target]",
            "NA: 40/45×100 ≈ 88.9. EU: 88.9. Worldwide: 80/45×100 ≈ 177.8.",
            "attainment is a measure function. vs_target op can emit a gap instead of a percent.",
        ],
    ]


def registry_rows(kind: str, rows: list[dict], extra_how: str) -> list[list]:
    out = []
    for r in rows:
        out.append(
            [
                r["name"],
                r["role"],
                r["enabled"],
                r["aliases"],
                r["description"],
                r["example"],
                extra_how,
                r["module"],
                r["attr"],
                r.get("min_args", ""),
                r.get("requires_value", ""),
            ]
        )
    return out


def all_operations_rows() -> list[list]:
    return [
        ["Aggregation", "sum", "Additive fold of a base. Densify fills missing periods with 0.", "base_measures.x: { sql: amount, agg: sum }", "Safe DuckDB GROUP BY + Pandas re-agg per cut"],
        ["Aggregation", "avg", "Weighted average. Carried as sum+count so a global avg is not mean-of-avgs.", "base_measures.x: { sql: amount, agg: avg }", "Recompute from __sum/__count at coarser cuts"],
        ["Aggregation", "count", "Row count. Densify fills 0.", "base_measures.n: { sql: order_id, agg: count }", "fill_zero like sum"],
        ["Row helper", "where:", "Pandas row mask before fold. Numeric ops coerce the compare column. ne excludes nulls (SQL-style).", "where: { column: mrr, op: gt, value: 0 }", "Ops: in eq ne gt gte lt lte between. between needs values: [lo, hi]. Not like/is_null."],
        ["Aggregation", "min / max", "Recomputed at every cut (not summed regional extrema).", "base_measures.peak: { sql: amount, agg: max }", ""],
        ["Aggregation", "count_distinct", "Distinct values. Non-additive; re-reads row-level detail.", "base_measures.n: { sql: order_id, agg: count_distinct }", "Not folded in DuckDB GROUP BY"],
        ["Aggregation", "median / percentile", "Non-additive. percentile requires percentile: 0–100.", "base_measures.p90: { sql: amount, agg: percentile, percentile: 90 }", "Row-level path"],
        ["Aggregation", "first / last", "Identity fold after sort by time (or window). Text columns are still coerced to numeric.", "base_measures.bal: { sql: balance, agg: last }", "Used with over: last_n"],
        ["Row helper", "expr", "Named row formula on columns or earlier helpers. No DuckDB SUM().", "gross: { expr: qty * price }", "Add agg: to fold, or identity_grain for point of"],
        ["Row helper", "lookup", "Map a column through a dict; default or strict.", "fee: { lookup: { column: pay, map: { COD: 25 }, default: 10 } }", ""],
        ["Row helper", "sql: column", "Pass a physical column (or simple ident). Default agg is sum if omitted.", "sotif_value: { sql: amount, agg: sum }", ""],
        ["Row helper", "columns + op", "Call a column function by name.", "ontime_full: { columns: [ontime, fullqty], op: multiply, agg: sum }", "See Column functions sheet"],
        ["Entity window", "over.lag / lead", "Value from another row in partition, ordered by order_by. Requires of:.", "prev: { over: { fn: lag, of: order_date, partition_by: [customer_id], order_by: [order_date] }, agg: max }", "Not calendar op: lag"],
        ["Entity window", "over.row_number / rank / dense_rank", "Sequence in partition. of: optional.", "seq: { over: { fn: row_number, partition_by: [customer_id], order_by: [order_date] }, agg: max }", ""],
        ["Entity window", "over.running_sum / running_avg", "Running aggregate along order_by. Requires of:.", "run: { over: { fn: running_sum, of: net, partition_by: [carrier_id], order_by: [start] }, agg: max }", ""],
        ["Entity window", "over.last_n", "JSON list of last n of values. Point only; cannot feed arithmetic.", "recent: { over: { fn: last_n, of: amount, n: 2, partition_by: [region], order_by: [order_id] }, agg: last }", "Dates as ISO strings"],
        ["Time", "grain day/week/month/quarter/year", "Bucket of the time column. Fiscal calendar only for Q/Y. Week is ISO-only.", "time: { column: event_month, grain: month, filter_code: reporting_month }", "parameters.time_grain overlay if declared"],
        ["Time", "periods year/quarter/month/week/day", "Independent predicates. Selection S = matching grain buckets; anchor = max(S). Point folds S; windows/trends from anchor.", "time: { periods: { year: year, month: \"current month\" } }", "Cannot set with compose:. Finer than grain is a bind error."],
        ["Time", "point offset", "Same measure over the (shifted) selection. Year-only + years:1 is the full prior year.", "previous_year_value: { of: sotif_value, op: point, offset: { years: 1 } }", ""],
        ["Time", "window trailing / leading", "Inclusive trailing N (or leading) periods from anchor as one number.", "value_3m: { of: sotif_value, op: window, trailing: { months: 3 }, inclusive: true }", "trailing: 3 from June is Apr-Jun"],
        ["Time", "window range", "ytd, qtd, mtd, wtd, cumulative, full_month/quarter/year. Named ranges are anchor-relative.", "value_qtd: { of: sotif_value, op: window, range: qtd }", "mtd under year=2026 is December"],
        ["Time", "trend", "Fixed-length array + trend_axes / trend_labels. Cap 50,000 cells/cut.", "trend_12m: { of: sotif_value, op: trend, trailing: { months: 12 } }", "Defaults to default_cut only"],
        ["Time", "snapshot (omit time:)", "No month filter; all matching rows. Host reporting_month skipped as no_time.", "Omit the time: block; only point offset 0 (and constants).", "No windows/trends/nonzero offsets"],
        ["Cuts", "group_by / ignore_filters / also_emit", "Named grains. G/R are examples, not hardcoded.", "cuts: [{ name: G, group_by: [], ignore_filters: [region], also_emit: [R] }]", "selected_dimensions overlays request grain"],
        ["Cuts", "output_cut / pack_also_emit", "Context output_cut is the also_emit walk root. YAML default does not lock. pack_also_emit: false emits only that cut.", "cuts: [{ name: G, also_emit: [R], pack_also_emit: false }]", "Declare parameters.output_cut to accept the overlay"],
        ["Cuts", "having", "Drop groups that fail measure predicates. Optional then_group_by.", "having: { predicates: [{ of: current_value, cmp: gt, value: 0 }] }", "cmp: gt gte lt lte eq ne between"],
        ["Cuts", "green_when", "Boolean green flag on survivors (above/below).", "green_when: { of: current_value, above: 40 }", "Does not drop rows"],
        ["Filters", "apply extract / calc / result", "DuckDB WHERE vs Pandas vs drop JSON rows.", "filters: { region: { apply: extract } }", "Unmapped valued filters are FilterError"],
        ["Multi-model", "model_relations", "Join two extracts on declared keys.", "model_relations: [{ left: a, right: b, on: [event_month, region], how: left }]", ""],
        ["Identity", "identity_grain", "Helpers as measures.of only when every emitted cut equals this grain.", "identity_grain: [carrier_id, shipment_id]", "Otherwise BindError; fold with agg:"],
        ["SQL model", "kind: sql", "Shape extract with CTEs; walked columns must be in output_schema.", "models/<id>.yaml kind: sql + sql: + output_schema:", "Not KPI sql: formulas"],
    ]


def pattern_rows() -> list[list]:
    return [
        [
            "Current month total",
            "Point of a summed base at offset 0",
            "base_measures:\n  sotif_value: { sql: amount, agg: sum }\nmeasures:\n  current_value:\n    of: sotif_value\n    op: point\n    offset: { months: 0 }",
            "op: point",
            "UI measure_key = current_value",
        ],
        [
            "Year-over-year %",
            "Two points + growth_pct (null/zero base → null)",
            "measures:\n  current_value: { of: sotif_value, op: point, offset: { months: 0 } }\n  previous_year_value: { of: sotif_value, op: point, offset: { years: 1 } }\n  yoy_month:\n    op: fn\n    fn: growth_pct\n    inputs: [current_value, previous_year_value]",
            "op: fn + fn: growth_pct  (alias yoy / mom)",
            "Same as op: arithmetic of: [current, previous] fn: divide only if you want a ratio, not %",
        ],
        [
            "Trailing 3-month sum",
            "Window inclusive of anchor",
            "measures:\n  value_3m:\n    of: sotif_value\n    op: window\n    trailing: { months: 3 }\n    inclusive: true",
            "op: window",
            "Do not sum three point measures by hand",
        ],
        [
            "QTD / YTD",
            "Named PTD range, not trailing 3",
            "measures:\n  value_qtd:\n    of: sotif_value\n    op: window\n    range: qtd",
            "op: window + range: qtd | ytd | mtd | wtd",
            "wtd needs time.grain: day",
        ],
        [
            "12-month graph",
            "Trend array; axis in response trend_axes",
            "measures:\n  trend_12m:\n    of: sotif_value\n    op: trend\n    trailing: { months: 12 }",
            "op: trend",
            "Restrict with cuts: [G] to avoid huge payloads",
        ],
        [
            "Ratio of two totals",
            "Per-measure expr after aggregation (not sum of row ratios)",
            "base_measures:\n  shipped_qty: { sql: shipped, agg: sum }\n  ordered_qty: { sql: ordered, agg: sum }\nmeasures:\n  shipped_now: { of: shipped_qty, op: point }\n  ordered_now: { of: ordered_qty, op: point }\n  fill_rate:\n    op: expr\n    expr: shipped_now / ordered_now",
            "op: expr  or  op: fn fn: divide",
            "If you wanted SUM(shipped/ordered) use base expr then agg: sum",
        ],
        [
            "Row qty × price then SUM",
            "Named row step, then fold",
            "base_measures:\n  gross: { expr: quantity * unit_price }\n  billed: { expr: gross, agg: sum }\nmeasures:\n  total_revenue: { of: billed, op: point }",
            "base expr + agg: sum",
            "lookup: for fee maps; CASE for shipping bands",
        ],
        [
            "Share of cut / rank",
            "percent_of_total is OVER(); fn: percent is this-row only",
            "measures:\n  group_share:\n    op: percent_of_total\n    of: current_value\n    partition_by: [region]\n  reason_rank:\n    op: rank\n    of: current_value\n    order: desc",
            "op: percent_of_total / rank / ntile / top_n",
            "partition_by ⊆ cut grain",
        ],
        [
            "Vs target / SLA flag",
            "constant + vs_target or threshold; predicate does not drop",
            "measures:\n  target: { op: constant, value: 95 }\n  gap:\n    op: vs_target\n    of: current_value\n    vs: target\n    as: gap\n  hit_sla:\n    op: threshold\n    of: current_value\n    cmp: gte\n    value: 95",
            "op: constant, vs_target, threshold, predicate",
            "having: drops groups; predicate keeps rows with 1/0",
        ],
        [
            "Same window last year",
            "Calendar lag of a window (not entity over:)",
            "measures:\n  value_3m: { of: sotif_value, op: window, trailing: { months: 3 } }\n  value_3m_ly:\n    op: lag\n    of: value_3m\n    offset: { years: 1 }",
            "op: lag / lead / diff / pct_change / index",
            "Cannot partition_by on calendar lag",
        ],
        [
            "Series stats (EWMA, hit rate)",
            "Hook over densified months; declare trailing so scan is wide enough",
            "measures:\n  smoothed:\n    op: hook\n    hook: ewma\n    of: sotif_value\n    trailing: { months: 12 }\n  months_on_sla:\n    op: hook\n    hook: hit_rate\n    of: sotif_value\n    trailing: { months: 12 }\n    value: 95",
            "op: hook",
            "See Hooks sheet; requires_value hooks need value:",
        ],
        [
            "Per-customer running billed",
            "over: on pre-fold detail, then identity or max fold",
            "base_measures:\n  running_net:\n    over:\n      fn: running_sum\n      of: net_charge\n      partition_by: [carrier_id]\n      order_by: [start, shipment_id]\n    agg: max",
            "over.fn: running_sum",
            "Caps: 500k rows, 50k partitions",
        ],
        [
            "Date math on columns named start/end",
            "date_diff in base expr; date_add as measure fn → ISO string",
            "base_measures:\n  transit_days: { expr: \"date_diff(start, end, 'day')\" }\n  delivered_on: { sql: end, agg: max }\nmeasures:\n  delivered_on: { of: delivered_on, op: point }\n  one: { op: constant, value: 7 }\n  next_review:\n    op: fn\n    fn: date_add\n    inputs: [delivered_on, one]\n    params: { unit: day }",
            "column/measure date_diff, date_add, epoch_day",
            "end is a legal column name outside CASE",
        ],
        [
            "Drop densified zeros",
            "HAVING after fill_zero",
            "having:\n  predicates:\n    - { of: current_value, cmp: gt, value: 0 }",
            "having",
            "Feb-only groups at a March pin become 0 and drop",
        ],
        [
            "Identity helper as UI column",
            "No agg:max workaround; identity_grain must match request grain",
            "identity_grain: [carrier_id, shipment_id]\nbase_measures:\n  net_charge: { expr: line_haul + fuel_amt }\nmeasures:\n  net_now: { of: net_charge, op: point, offset: { months: 0 } }",
            "identity_grain + op: point of helper",
            "Coarser also_emit cut → BindError",
        ],
        [
            "Attainment vs goal",
            "actual / target × 100",
            "measures:\n  vs_goal:\n    op: fn\n    fn: attainment\n    inputs: [current_value, target]",
            "fn: attainment",
            "Zero/null target → null",
        ],
    ]


def extend_rows() -> list[list]:
    return [
        [
            "New KPI (existing catalog)",
            "Authoring",
            "Copy kpi_config/kpis/sotif/3004.yaml → <kpi_group>/<kpi_id>.yaml. Fill time, dimensions, base_measures, cuts, measures. Host module_path stays kpi_engine.main.",
            "kpi_config/kpis/<kpi_group>/<id>.yaml",
            "Do not add a per-KPI Python module. Do not edit pipeline/.",
        ],
        [
            "New extract / joins",
            "Model",
            "Add kpi_config/models/<kpi_group>/<id>.yaml (kind: physical or sql). Point KPI model: at model_id. Paths stay on context.datasets.",
            "kpi_config/models/<kpi_group>/<id>.yaml",
            "KPI formulas stay catalog ops; messy SQL belongs in the model.",
        ],
        [
            "New measure op kind",
            "Measure ops",
            "1) Implement OpPlugin in capabilities/ops/ (combo, cut, or period). 2) Add key in registries/ops.yaml (role, description, example, module, attr). 3) Regenerate CAPABILITIES.md.",
            "capabilities/ops/*.py + registries/ops.yaml",
            "Not a pipeline/ edit. KPI YAML uses op: <new_name>.",
        ],
        [
            "New row function",
            "Column functions",
            "1) Add function in capabilities/functions/column/impl.py taking pandas Series. 2) Register in registries/functions/column.yaml. 3) Use base_measures columns:+op: or expr: name(…). 4) Regen CAPABILITIES.md.",
            "column/impl.py + functions/column.yaml",
            "min_args if variadic. Aliases optional.",
        ],
        [
            "New scalar function",
            "Measure functions",
            "1) Add function in capabilities/functions/measure/impl.py. 2) Register in registries/functions/measure.yaml. 3) Call via op: fn / arithmetic / expr. 4) Regen CAPABILITIES.md.",
            "measure/impl.py + functions/measure.yaml",
            "Reject lists/trends if the fn is scalar-only (see date_diff).",
        ],
        [
            "New series hook",
            "Hooks",
            "1) Add function in capabilities/hooks/impl.py (densified series in, scalar out). 2) registries/hooks.yaml with requires_value / extra_keys if needed. 3) measures op: hook + hook: name + trailing/offset. 4) Regen CAPABILITIES.md.",
            "hooks/impl.py + hooks.yaml",
            "Must not open DuckDB or read ADLS. Planner uses trailing to widen scan.",
        ],
        [
            "New aggregation",
            "Engine (rare)",
            "contracts.py + binder + calc_engine fold rules. This is not a catalog name.",
            "kpi_engine/pipeline/ + contracts.py",
            "Only when sum/avg/count/… cannot express the fold.",
        ],
        [
            "New filter operator",
            "Engine (rare)",
            "filter_ops.py allowlist + DuckDB/Pandas impl.",
            "pipeline/filter_ops.py",
            "Default host operator is IN.",
        ],
        [
            "Disable a name",
            "Registry",
            "Set enabled: false in the YAML registry. Bind will reject it.",
            "registries/*.yaml",
            "Do not delete platform names without a migration.",
        ],
        [
            "Regenerate docs",
            "Catalog",
            "After any registry change, regenerate kpi_engine/registries/CAPABILITIES.md (write_generated_docs).",
            "registries/CAPABILITIES.md",
            "Do not hand-edit CAPABILITIES.md.",
        ],
    ]


def main() -> None:
    OUT.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    cover(wb)

    write_sheet(
        wb,
        "Framework components",
        ["Component", "One-line meaning", "What it is", "YAML / where", "What it is not", "When to use"],
        component_rows(),
        [28, 42, 62, 48, 42, 48],
    )
    write_sheet(
        wb,
        "Naming conventions",
        ["Topic", "Convention", "Do this", "Do not", "Engine rule"],
        convention_rows(),
        [28, 58, 52, 48, 58],
    )
    write_sheet(
        wb,
        "YAML preparation",
        ["Topic", "What to provide / emit", "Do this", "Do not", "Engine / AI rule"],
        preparation_rows(),
        [28, 58, 62, 48, 58],
    )
    write_sheet(
        wb,
        "Decision guide",
        ["Step", "If you need…", "Use this", "Do not use", "Why / note"],
        decision_rows(),
        [8, 58, 48, 42, 48],
    )
    write_sheet(
        wb,
        "Sample dataset",
        ["event_month", "region", "order_id", "qty", "price", "shipped", "amount (qty×price)", "Note"],
        sample_fact_rows(),
        [14, 12, 12, 10, 10, 12, 22, 42],
    )
    write_sheet(
        wb,
        "Sample monthly totals",
        ["period", "grain", "revenue", "How it was added"],
        sample_monthly_rows(),
        [14, 16, 12, 40],
    )
    write_sheet(
        wb,
        "Worked examples",
        ["Example", "Components used", "Business question", "YAML", "Arithmetic on the sample (anchor Mar 2026)", "Takeaway"],
        worked_example_rows(),
        [36, 36, 42, 62, 52, 48],
    )

    headers = [
        "Name",
        "Role",
        "Enabled",
        "Aliases",
        "What it means",
        "How to use in YAML",
        "When to use / notes",
        "Python module",
        "Attr",
        "min_args",
        "requires_value",
    ]
    widths = [22, 12, 10, 28, 48, 52, 40, 42, 22, 12, 16]

    ops = load_registry("ops.yaml")
    write_sheet(
        wb,
        "Measure ops",
        headers,
        registry_rows(
            "op",
            ops,
            "Put under measures: as op: <name>. Combo ops (point/window/fn/…) run per group; "
            "cut ops (rank/share/…) need the full cut; period ops (lag/index/…) need time:.",
        ),
        widths,
    )

    cols = load_registry("functions/column.yaml")
    write_sheet(
        wb,
        "Column functions",
        headers,
        registry_rows(
            "col",
            cols,
            "Use in base_measures with columns: + op:, or inside expr:/sql: as name(...). "
            "Then fold with agg:. These run per retrieved row, not on totals.",
        ),
        widths,
    )

    meas = load_registry("functions/measure.yaml")
    write_sheet(
        wb,
        "Measure functions",
        headers,
        registry_rows(
            "meas",
            meas,
            "Use under measures with op: fn (fn: + inputs:), op: arithmetic, or op: expr. "
            "Inputs are other measure keys (already aggregated). Not physical columns.",
        ),
        widths,
    )

    hooks = load_registry("hooks.yaml")
    write_sheet(
        wb,
        "Hooks",
        headers,
        registry_rows(
            "hook",
            hooks,
            "measures: { op: hook, hook: <name>, of: <base>, trailing: { months: N } }. "
            "Declare trailing/offset so required_span is wide enough. Hooks that list requires_value need value:.",
        ),
        widths,
    )

    write_sheet(
        wb,
        "All operations",
        ["Category", "Operation", "What it means", "YAML how-to", "Notes"],
        all_operations_rows(),
        [18, 32, 62, 72, 42],
    )

    write_sheet(
        wb,
        "YAML calculation patterns",
        ["Pattern", "Intent", "YAML example", "Catalog names used", "Pitfall to avoid"],
        pattern_rows(),
        [28, 42, 70, 40, 48],
    )

    write_sheet(
        wb,
        "How to extend",
        ["Goal", "Capability type", "Steps", "Files to change", "Do not"],
        extend_rows(),
        [32, 20, 78, 42, 42],
    )

    wb.save(OUT)
    print(f"Wrote {OUT}  ops={len(ops)} column_fns={len(cols)} measure_fns={len(meas)} hooks={len(hooks)}")


if __name__ == "__main__":
    main()
